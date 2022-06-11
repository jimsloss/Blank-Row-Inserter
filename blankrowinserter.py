# This program takes two numbers (N and M) and filename string from the command
# line. Starting at row N the program inserts M blank rows in to the spreadsheet
# called filename

def blankRowInserter():

    # need sys to read from command line,
    # need openpyxl to work with excel
    # need os for excel file location check

    import sys, openpyxl, os
    from openpyxl.styles import Font, Alignment

    # get arguments from the command line
    
    if len(sys.argv) == 4: # check that there are 3 arguments

        N = sys.argv[1]    # set N to the 1st value given
        M = sys.argv[2]    # set M to the 2nd value given
        filename = sys.argv[3]  # set filename to the name provided

        while not(N.isnumeric()):
            N = input("Which row number shall the blank rows start at? :")

        N = int(N)

        while not(M.isnumeric()):
            M = input("How many blank rows are to be inserted? :")

        M = int(M)

        while not(type(filename) == str):
            filename = input("What is the name of the spreedsheet file? :")

    else:
        print("You need to provide 2 numbers and a string to  continue")
        print("Program terminated - try again")
        sys.exit()


   # check if filename exists

    if os.path.exists(filename):
        # wb = openpyxl.Workbook()
        # sheet = wb.active
        wb= openpyxl.load_workbook(filename)
        print("Spreadsheet found")

        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        sheet.insert_rows(N,M)
        wb.save(filename)
        print("done, rows inserted")

    else:
        print("no file by that name in the current directly")
        print("program terminated. Try again.")

        
    

blankRowInserter()
        
